import cv2 as cv
from openpyxl import Workbook
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import numpy as np
import pylab as pl
import threading
import sys
from tkinter.filedialog import askopenfilename
global filename
global finished
global xx1
global xx2
global yy1
global yy2
global cars_detected
global seconds
global paths
global kill

# Initialize the parameters
confThreshold = 0.5  # Confidence threshold
nmsThreshold = 0.4  # Non-maximum suppression threshold
inpWidth = 416  # Width of network's input image
inpHeight = 416  # Height of network's input image

# Load names of classes
classesFile = "coco.names"
classes = None
with open(classesFile, 'rt') as f:
    classes = f.read().rstrip('\n').split('\n')

# Give the configuration and weight files for the model and load the network using them.
modelConfiguration = "yolov3.cfg"
modelWeights = "yolov3.weights"

net = cv.dnn.readNetFromDarknet(modelConfiguration, modelWeights)
net.setPreferableBackend(cv.dnn.DNN_BACKEND_OPENCV)
net.setPreferableTarget(cv.dnn.DNN_TARGET_CPU)



# Get the names of the output layers
def getOutputsNames(net):
    # Get the names of all the layers in the network
    layersNames = net.getLayerNames()
    # Get the names of the output layers, i.e. the layers with unconnected outputs
    return [layersNames[i[0] - 1] for i in net.getUnconnectedOutLayers()]


# Draw the predicted bounding box
def ClassDetect(classId, left, top, right, bottom, rz, cadr):
    global count
    global narush
    global xx1
    global xx2
    global yy1
    global yy2
    global paths
    global seconds
    if (classes[classId] == "car") or (classes[classId] == "person"):
        if (classes[classId] == "car"):
            return 1
        else:
            cenx = (left + right) / 2
            ceny = (top + bottom) / 2
            for i in range(len(xx1)-1):
                k = (yy2[i] - yy1[i]) / (xx2[i] - xx1[i])
                b = (xx2[i] * yy1[i] - xx1[i] * yy2[i]) / (xx2[i] - xx1[i])
                if (cenx * k + b > ceny - 5 and cenx*k+b < ceny + 5) and (cadr % 50 == 0 or cadr == 0):
                    cv.rectangle(rz, (left, top), (right, bottom), (255, 178, 50), 3)
                    cv.imshow("Anomaly", rz)
                    path = "data/" + str(narush) + ".png"
                    paths.append(path)
                    seconds.append(cadr % 25)
                    cv.imwrite(path, rz)
                    narush += 1
                    messagebox.showinfo("Anomaly", "Нарушение ПДД!")
                    cv.waitKey(3000)
            return 2
    else:
        return 0


# Remove the bounding boxes with low confidence using non-maxima suppression
def postprocess(rz, outs, rect, cadr):
    frameHeight = rz.shape[0]
    frameWidth = rz.shape[1]
    # Scan through all the bounding boxes output from the network and keep only the
    # ones with high confidence scores. Assign the box's class label as the class with the highest score.
    classIds = []
    confidences = []
    boxes = []
    cadrr = cadr
    for out in outs:
        for detection in out:
            scores = detection[5:]
            classId = np.argmax(scores)
            confidence = scores[classId]
            if confidence > confThreshold:
                center_x = int(detection[0] * frameWidth)
                center_y = int(detection[1] * frameHeight)
                width = int(detection[2] * frameWidth)
                height = int(detection[3] * frameHeight)
                left = int(center_x - width / 2)
                top = int(center_y - height / 2)
                classIds.append(classId)
                confidences.append(float(confidence))
                boxes.append([left, top, width, height])

    # Perform non maximum suppression to eliminate redundant overlapping boxes with
    # lower confidences.
    indices = cv.dnn.NMSBoxes(boxes, confidences, confThreshold, nmsThreshold)
    current = 0
    if rect == -1:
        for i in indices:
            i = i[0]
            box = boxes[i]
            left = box[0]
            top = box[1]
            width = box[2]
            height = box[3]
            if ClassDetect(classIds[i], left, top, left + width, top + height, rz, cadrr) == 2:
                return box
            else:
                return -1
    elif len(rect) > 1:
        sq = 0
        for i in indices:
            i = i[0]
            box = boxes[i]
            center_x = box[0] + box[2] / 2
            center_y = box[1] + box[3] / 2
            left = box[0]
            top = box[1]
            width = box[2]
            height = box[3]
            if ((rect[0] < center_x) and (center_x < (rect[0] + rect[2])) and (rect[1] < center_y) and
                (center_y < (rect[1] + rect[3]))) and \
                    ClassDetect(classIds[i], left, top, left + width, top + height, rz, cadrr) == 1:
                current += 1
                sqr_box = (box[2] * box[3])
                sq += sqr_box
        if current == 0:
            x = rect[0]
            y = rect[1]
            w = rect[2]
            h = rect[3]
            cv.rectangle(rz, (int(x), int(y)), (int(x + w), int(y + h)), (255, 178, 50), 3)
            cv.waitKey(1000)
        return sq, current


def openfile():
    name = askopenfilename(initialdir="C:/",
                           filetypes=(("Видео MP4", "*.mp4"), ("Видео AVI", ".avi"), ("Все файлы", "*.*")),
                           title="Выберите файл."
                           )

    global filename
    if (name == "") or (not (name[-4:] == ".mp4") and not (name[-4:] == ".avi")):
        filename = 0
    else:
        tishka.configure(text="Выбранный файл = " + name, bg="#d1e1fc")
        jams.configure(state=NORMAL)
        PDD.configure(state=NORMAL)
        stealing.configure(state=NORMAL)
        excel_button.configure(state=NORMAL)
        graph.configure(state=NORMAL)
        filename = name


def exiting():
    global kill
    if messagebox.askokcancel("Выход", "Вы действительно хотите выйти?"):
        kill = True
        root.destroy()
        sys.exit()


def plot():
    global cars_detected
    pl.subplot(211)
    pl.gcf().canvas.set_window_title('Car detect')
    pl.plot(cars_detected, label="Cars detected")
    pl.xlabel('time, sec')
    pl.ylabel('cars')
    pl.legend()
    pl.plot(cars_detected)
    pl.show()


def excel():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Num"
    ws['B1'] = "Sec"
    ws['C1'] = "Path"
    for i in range(len(paths)):
        ws.cell(row=i + 2, column=1).value = i + 1
        ws.cell(row=i + 2, column=2).value = seconds[i]
        ws.cell(row=i + 2, column=3).value = paths[i]
    wb.save('base.xlsx')


def lines(rz, count):
    linelenght = 50
    if (count == 20) or (len(xx1) < 1):
        count = 20
        while (count > 14):
            xx1.clear()
            xx2.clear()
            yy1.clear()
            yy2.clear()
            count = 0
            frame = cv.GaussianBlur(rz, (5, 5), 0)
            gray = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
            edges = cv.Canny(gray, 75, 150)
            lines = cv.HoughLinesP(edges, 1, np.pi / 180, linelenght, maxLineGap=250)
            if lines is not None:
                for line in lines:
                    x1, y1, x2, y2 = line[0]
                    xx1.append(x1)
                    xx2.append(x2)
                    yy1.append(y1)
                    yy2.append(y2)
                    count = count + 1
            if (count > 10) or (count < 1):
                if count > 14:
                    linelenght = linelenght + 1
                else:
                    linelenght = linelenght - 1
    return count


def detection(toggle):
    global xx1
    global xx2
    global yy1
    global yy2
    global count
    global finished
    global filename
    global vid_writer
    if filename is None:
        tishka.configure(text="Choose file")
        return -1
    elif filename == 0:
        tishka.configure(text="Wrong file")
        return -1
    else:
        name = filename
        cap = cv.VideoCapture(name)
        hasFrame, frame = cap.read()
        cv.imwrite("frame.jpg", frame)
        img = cv.imread("frame.jpg")
        scale_percent = 30  # percent of original size
        width = int(img.shape[1] * scale_percent / 100)
        height = int(img.shape[0] * scale_percent / 100)
        dim = (width, height)
        cv.destroyAllWindows()
        cadr = 0
        count = 20
        xx1 = []
        xx2 = []
        yy1 = []
        yy2 = []
        thread = threading.Thread(target=processing)
        thread.start()
        count = 20
        while cv.waitKey(1) < 0:
            if kill:
                break
            # get frame from the video
            hasFrame, frame = cap.read()
            # Stop the program if reached end of video
            if not hasFrame:
                cv.waitKey(3000)
                # Release device
                cap.release()
                break
            # Create a 4D blob from a frame.
            rz = cv.resize(frame, dim, interpolation=cv.INTER_AREA)
            blob = cv.dnn.blobFromImage(rz, 1 / 255, (inpWidth, inpHeight), [0, 0, 0], 1, crop=False)

            # Sets the input to the network
            net.setInput(blob)

            # Runs the forward pass to get output of the output layers
            outs = net.forward(getOutputsNames(net))

            # Remove the bounding boxes with low confidence
            count = lines(rz, count)
            bbox = postprocess(rz, outs, -1, cadr)
            cadr +=1
        finished = True
        cap.release()
        if toggle == 0:
            if narush == 0:
                messagebox.showinfo("Информация", "Нарушителей обнаружено не было")
            else:
                messagebox.showwarning("Информация", "Было обнаружено " + str(narush) + " нарушений")
            cv.destroyAllWindows()
        if toggle == 1:
            excel()


def resize_frame(frame):
    shape = (768, 1024)
    width = int(frame.shape[1] * (((shape[1] / frame.shape[1]) * 100)) / 100)
    height = int(frame.shape[0] * (((shape[0] / frame.shape[0]) * 100)) / 100)
    dim = (width, height)
    # resize image
    resized = cv.resize(frame, dim, interpolation=cv.INTER_AREA)
    return resized


def resize_rect(r, frame):
    shape = (768, 1024)
    width = frame.shape[1] / shape[1]
    height = frame.shape[0] / shape[0]
    rect = (r[0] * width, r[1] * height, r[2] * width, r[3] * height)
    return rect


def processing():
    global finished
    finished = False
    ph = 1
    dir = 1
    while not finished:
        if kill:
            break
        if ph == 10:
            ph = 1
            dir = dir + 1
        if dir == 3:
            dir = 1
        pic = str(dir) + "/" + str(ph) + ".jpg"
        imgFile = cv.imread(pic)
        cv.imshow("Processing", imgFile)
        cv.waitKey(1000)
        ph = ph + 1


def select_rect(frame):
    messagebox.showinfo("Tutorial",
                        "Поставьте курсор на точку начала прямоугольника.\nНажмите и удерживайте правую кнопку мыши. \nПеретащите курсор на точку конца прямоугольника и отпустите правую кнопку. Нажмите Enter.")
    resize_img = resize_frame(frame)
    rect = cv.selectROI("Image", resize_img, False, False)
    cv.destroyAllWindows()
    return rect


def traffic_jam(mode, toggle):
    global xx1
    global xx2
    global yy1
    global yy2
    global cars_detected
    global finished
    global filename
    thread = threading.Thread(target=processing)
    average = 0
    count = 20
    cap = cv.VideoCapture(filename)
    hasFrame, frame = cap.read()
    if (mode == 0):
        r = select_rect(frame)
        cv.destroyAllWindows()
        rect = resize_rect(r, frame)
        sqr_rect = (rect[2] * rect[3])
        thread.start()

    if (mode == 1):
        thread.start()
        count = 20
        while count > 14 or count < 1:
            hasFrame, frame = cap.read()
            if (hasFrame == False):
                break
            count = lines(frame, count)
        sr = 0
        i = 0
        for i in range(len(xx1)):
            sr += xx1[i]
            sr += xx2[i]
        i *= 2
        sr /= i
        min_height_left = frame.shape[0]
        max_height_left = 0
        min_height_right = frame.shape[0]
        max_height_right = 0

        for i in range(len(yy1)):
            if (xx1[i] < sr):
                if (yy1[i] < min_height_left):
                    min_height_left = yy1[i]
                    min_height_left_index = i
                elif (yy1[i] > max_height_left):
                    max_height_left = yy1[i]
                    max_height_left_index = i
            else:
                if (yy1[i] < min_height_right):
                    min_height_right = yy1[i]
                    min_height_right_index = i
                elif (yy1[i] > max_height_right):
                    max_height_right = yy1[i]
                    max_height_right_index = i
            if (xx2[i] < sr):
                if (yy2[i] < min_height_left):
                    min_height_left = yy2[i]
                    min_height_left_index = i
                elif (yy2[i] > max_height_left):
                    max_height_left = yy2[i]
                    max_height_left_index = i
            else:
                if (yy2[i] < min_height_right):
                    min_height_right = yy2[i]
                    min_height_right_index = i
                elif (yy2[i] > max_height_right):
                    max_height_right = yy2[i]
                    max_height_right_index = i
        if (yy1[min_height_left_index] == min_height_left):
            x_min_left = xx1[min_height_left_index]
        else:
            x_min_left = xx2[min_height_left_index]

        if (yy1[max_height_right_index] == max_height_right):
            x_max_right = xx1[max_height_right_index]
        else:
            x_max_right = xx2[max_height_right_index]

        sqr_min_max = (x_max_right - x_min_left) * (max_height_right - min_height_left)

        if (yy1[max_height_left_index] == max_height_left):
            x1_min_left = xx1[max_height_left_index]
        else:
            x1_min_left = xx2[max_height_left_index]

        if (yy1[min_height_right_index] == min_height_right):
            x1_max_right = xx1[min_height_right_index]
        else:
            x1_max_right = xx2[min_height_right_index]

        sqr_max_min = (x1_max_right - x1_min_left) * (max_height_left - min_height_right)

        sqr_rect = (sqr_min_max + sqr_max_min) / 2

        if (sqr_min_max > sqr_max_min):
            rect = (
                x_min_left, min_height_left, (x_max_right - x_min_left),
                (max_height_right - min_height_left))  # x,y,w,h
        else:
            rect = (x1_min_left, min_height_right, (x1_max_right - x1_min_left), (max_height_left - min_height_right))

        img_rect = cv.rectangle(frame, (int(rect[0]), int(rect[1])), (int(rect[0] + rect[2]), int(rect[1] + rect[3])),
                                (255, 178, 50), 3)

    cadr = 0
    ret = 0
    sum_relative = 0
    pred = 0
    pysto = 0
    probka = 0

    while (ret != -1):
        if kill:
            break
        hasFrame, frame = cap.read()
        if not hasFrame:
            break

        blob = cv.dnn.blobFromImage(frame, 1 / 255, (inpWidth, inpHeight), [0, 0, 0], 1, crop=False)

        # Sets the input to the network
        net.setInput(blob)

        # Runs the forward pass to get output of the output layers
        outs = net.forward(getOutputsNames(net))
        sq, ret = postprocess(frame, outs, rect, cadr)
        cars_detected.append(ret)
        sum_relative += (sq / sqr_rect)
        if (ret != -1):
            if (pred == ret):
                if (ret == 0):
                    pysto += 1
                else:
                    probka += 1

        pred = ret
        cadr += 1
        if ((cadr % 50) == 0):
            sum_relative /= 50
            probka /= 50
            coef_prob = (sum_relative + probka) / 2
            average += int(coef_prob * 10)
            count += 1
            f = resize_frame(frame)
            label = "Ball: %.0f" % (coef_prob * 10)
            if coef_prob > 0.7:
                cv.rectangle(f, (0, 0),
                             (260, 53), (255, 255, 255), cv.FILLED)
                cv.putText(f, label, (5, 50), cv.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 3)
                cv.destroyAllWindows()
                cv.imshow("Warning", f)
                cv.waitKey(1000)
            probka = 0
            sum_relative = 0
    messagebox.showinfo("INFO", "Средний балл пробки: " + str(int(average / count)))
    finished = True
    if toggle == 1:
        plot()
    jams_thread.join()


def steal():
    global finished
    global filenam
    steal = 0
    ret = 0
    cadr = 0
    cap = cv.VideoCapture(filename)
    hasFrame, frame = cap.read()
    r = select_rect(frame)

    cv.destroyAllWindows()
    rect = resize_rect(r, frame)
    rect1 = rect

    tracker = cv.TrackerMOSSE_create()
    ok = tracker.init(frame, rect)

    thread = threading.Thread(target=processing)
    thread.start()

    while (ret != -1):
        if kill:
            break
        hasFrame, frame = cap.read()
        if not hasFrame:
            break
        rz = resize_frame(frame)
        hasFrame, rect = tracker.update(frame)

        if hasFrame:

            if (((rect1[0]) < (rect[0] + rect[2] / 2)) and ((rect[0] + rect[2] / 2) < (rect1[0] + rect1[2])) and (
                    rect1[1] < (rect[1] + rect[3] / 2)) and ((rect[1] + rect[3] / 2) < (rect1[1] + rect1[3]))):
                steal = 0
            else:
                steal += 1

        if (steal == 25):
            break
        cadr += 1

    finished = True
    if (steal == 25):
        cv.rectangle(frame, (int(rect[0]), int(rect[1])), (int(rect[0] + rect[2]), int(rect[1] + rect[3])), (0, 0, 255),
                     3)
        # cv.rectangle(rz, (int(x), int(y)), (int(x + w), int(y + h)), (255, 178, 50), 3)
        img = resize_frame(frame)
        messagebox.showerror("ALERT", "Украли машину")
        cv.imshow("ALERT", img)
        cv.waitKey(0)


def choose(mode):
    toggle = messagebox.askyesno("Question", "Вы хотите выбрать дорогу в автоматическом режиме?\n" +
                                 "Автоматическое определение находится в тестовом режиме")
    if mode == 0:
        if toggle == True:
            traffic_jam(1, 0)
        else:
            traffic_jam(0, 0)
    else:
        if toggle == True:
            traffic_jam(1, 1)
        else:
            traffic_jam(0, 1)

kill = False
cars_detected = []
xx1 = []
xx2 = []
yy1 = []
yy2 = []
seconds = []
paths = []
filename = None
count = 0
finished = None
narush = 0
root = Tk()
root.iconbitmap("cat.ico")
root.resizable(width=False, height=False)
root.geometry("800x600")
root.title('Stage 2')
jams_thread = threading.Thread(target=lambda: choose(0))
PDD_thread = threading.Thread(target=lambda: detection(0))
stealing_thread = threading.Thread(target=lambda: steal())
graph_thread = threading.Thread(target=lambda: choose(1))
excel_thread = threading.Thread(target=lambda: detection(1))
photo = PhotoImage(file="cat.gif")
label = ttk.Label(root, image=photo)
PhotoImage(file="cat.gif")
label.pack()
tishka = Label(root, text="", bg="#cae2f9", font=("Arial Bold", 15))
tishka.place(x=30, y=555)
choose_button = tk.Button(root, bg='#fff7ea', fg='black', relief=RAISED, text='Выбрать видео', font=("Arial Bold", 9),
                          width=22, height=2, command=openfile)
jams = tk.Button(root, bg='#fff7ea', text="Анализ видеофрагмента\nна наличие пробок", font=("Arial Bold", 9), width=22,
                 height=2, relief=RAISED, fg="black", command=lambda: jams_thread.start(), state=DISABLED)
exit_button = tk.Button(root, bg='#fff7ea', text="Выход", font=("Arial Bold", 9), width=22, height=2, relief=RAISED,
                        fg="black", command=exiting)
PDD = tk.Button(root, bg='#fff7ea', text="Обнаружение пешеходов\nнарушителей ПДД", font=("Arial Bold", 9), width=22,
                height=2, relief=RAISED, fg="black", command=lambda: PDD_thread.start(), state=DISABLED)
stealing = tk.Button(root, bg='#fff7ea', text="Выявление угона\nавтомобиля", font=("Arial Bold", 9), width=22, height=2,
                     relief=RAISED, fg="black", command=lambda: stealing_thread.start(), state=DISABLED)
excel_button = tk.Button(root, bg='#fff7ea', text="Запись в таблицу данных\nо нарушителях", font=("Arial Bold", 9),
                         width=22, height=2, relief=RAISED, fg="black", command=lambda: excel_thread.start(), state=DISABLED)
graph = tk.Button(root, bg='#fff7ea', text="Постоение графика потока\n", font=("Arial Bold", 9), width=22, height=2,
                  relief=RAISED, fg="black", command=lambda: graph_thread.stars(), state=DISABLED)
choose_button.place(x=122, y=150)
jams.place(x=24, y=290)
exit_button.place(x=120, y=425)
PDD.place(x=24, y=220)
stealing.place(x=121, y=360)
excel_button.place(x=217, y=220)
graph.place(x=218, y=290)
root.protocol("WM_DELETE_WINDOW", exiting)
root.mainloop()
